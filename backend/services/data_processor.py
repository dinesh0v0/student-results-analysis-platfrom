import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Tuple

from fastapi import UploadFile
from openpyxl import load_workbook


MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_UPLOAD_ROWS = 5000
DEFAULT_MAX_MARKS = 100.0
PASS_PERCENTAGE = 40.0
REQUIRED_COLUMNS = [
    "register_number",
    "student_name",
    "semester",
    "subject_code",
    "marks",
]
TEXT_LIMITS = {
    "campus": 255,
    "faculty": 255,
    "department": 255,
    "branch": 255,
    "section": 100,
    "register_number": 50,
    "student_name": 255,
    "subject_code": 50,
    "subject_name": 255,
}
VALID_GRADES = {
    "O",
    "A+",
    "A",
    "B+",
    "B",
    "C",
    "F",
    "P",
    "PASS",
    "FAIL",
    "AB",
    "N/A",
}


COLUMN_ALIASES = {
    "campus": ["campus"],
    "faculty": ["faculty", "school"],
    "department": ["department", "dept", "department_name"],
    "branch": ["branch", "program", "course", "specialization"],
    "section": ["section", "class_section", "group"],
    "register_number": [
        "register_number",
        "register no",
        "register_no",
        "reg_no",
        "reg no",
        "roll_number",
        "roll no",
        "rollno",
        "registration_number",
        "enrollment_no",
    ],
    "student_name": [
        "student_name",
        "student name",
        "name",
        "full_name",
        "full name",
    ],
    "semester": ["semester", "sem", "semester_no", "sem_no"],
    "subject_code": [
        "subject_code",
        "subject_codes",
        "subject code",
        "sub_code",
        "sub code",
        "course_code",
    ],
    "subject_name": [
        "subject_name",
        "subject_names",
        "subject name",
        "sub_name",
        "sub name",
        "course_name",
        "course name",
    ],
    "marks": [
        "marks",
        "marks_obtained",
        "marks obtained",
        "score",
        "obtained_marks",
    ],
    "max_marks": [
        "max_marks",
        "max marks",
        "maximum_marks",
        "out_of",
        "out of",
        "total",
    ],
    "grade": ["grade", "grades", "letter_grade", "letter grade"],
}


def _normalize_column_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _build_alias_lookup() -> Dict[str, str]:
    alias_lookup: Dict[str, str] = {}
    for standard_name, aliases in COLUMN_ALIASES.items():
        alias_lookup[_normalize_column_name(standard_name)] = standard_name
        for alias in aliases:
            alias_lookup[_normalize_column_name(alias)] = standard_name
    return alias_lookup


ALIAS_LOOKUP = _build_alias_lookup()


def _resolve_headers(raw_headers: List[Any]) -> Tuple[List[str], List[str]]:
    normalized_headers = [_normalize_column_name(header) for header in raw_headers]
    resolved_headers: List[str] = []
    seen_headers: set[str] = set()
    errors: List[str] = []

    if not any(normalized_headers):
        return [], ["The uploaded file must include a valid header row."]

    for header in normalized_headers:
        if not header:
            errors.append("The uploaded file contains an empty column header.")
            continue

        standard_name = ALIAS_LOOKUP.get(header, header)
        if standard_name in seen_headers:
            errors.append(f"Duplicate column detected: '{standard_name}'.")
            continue

        resolved_headers.append(standard_name)
        seen_headers.add(standard_name)

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in seen_headers]
    if missing_columns:
        errors.append(
            "Missing required columns: " + ", ".join(missing_columns) + "."
        )

    return resolved_headers, errors


def _clean_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ").strip()
    if text.lower() in {"", "nan", "none"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _read_csv_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        decoded_content = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return [], [
            "Unable to read the CSV file. Please save it as UTF-8 encoded CSV and try again."
        ]

    reader = csv.DictReader(io.StringIO(decoded_content))
    if reader.fieldnames is None:
        return [], ["The uploaded CSV file must include a header row."]

    headers, header_errors = _resolve_headers(list(reader.fieldnames))
    if header_errors:
        return [], header_errors

    reader.fieldnames = headers
    rows: List[Dict[str, Any]] = []

    for row_number, raw_row in enumerate(reader, start=2):
        if None in raw_row and any(_clean_text(value) for value in raw_row.get(None, [])):
            return [], [f"Row {row_number} has more values than the header row."]

        rows.append(dict(raw_row))
        if len(rows) > MAX_UPLOAD_ROWS:
            return [], [
                f"The uploaded file exceeds the {MAX_UPLOAD_ROWS} row limit. Please split it into smaller files."
            ]

    return rows, []


def _read_xlsx_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        row_iterator = sheet.iter_rows(values_only=True)
        header_row = next(row_iterator, None)
        if header_row is None:
            return [], ["The uploaded file is empty."]

        headers, header_errors = _resolve_headers(list(header_row))
        if header_errors:
            return [], header_errors

        rows: List[Dict[str, Any]] = []
        header_count = len(headers)

        for row_number, values in enumerate(row_iterator, start=2):
            values = tuple(values or ())
            trailing_values = values[header_count:]
            if any(_clean_text(value) for value in trailing_values):
                return [], [f"Row {row_number} has more values than the header row."]

            row: Dict[str, Any] = {}
            for index, header in enumerate(headers):
                row[header] = values[index] if index < len(values) else None

            rows.append(row)
            if len(rows) > MAX_UPLOAD_ROWS:
                return [], [
                    f"The uploaded file exceeds the {MAX_UPLOAD_ROWS} row limit. Please split it into smaller files."
                ]

        return rows, []
    except Exception:
        return [], ["Unable to read the Excel file. Please upload a valid .xlsx file."]
    finally:
        if workbook is not None:
            workbook.close()


def _parse_semester(value: Any, row_number: int) -> Tuple[int | None, str | None]:
    text = _clean_text(value)
    if not text:
        return None, f"Row {row_number} is missing 'semester'."

    try:
        decimal_value = Decimal(text)
    except InvalidOperation:
        return None, f"Row {row_number} has an invalid semester value '{text}'."

    if decimal_value != int(decimal_value):
        return None, f"Row {row_number} has an invalid semester value '{text}'."

    semester = int(decimal_value)
    if semester < 1 or semester > 12:
        return None, f"Row {row_number} must have a semester between 1 and 12."

    return semester, None


def _parse_number(
    value: Any,
    row_number: int,
    field_name: str,
    *,
    allow_blank: bool = False,
    minimum: float | None = None,
) -> Tuple[float | None, str | None]:
    text = _clean_text(value)
    if not text:
        if allow_blank:
            return None, None
        return None, f"Row {row_number} is missing '{field_name}'."

    try:
        number = float(Decimal(text))
    except InvalidOperation:
        return None, f"Row {row_number} has an invalid {field_name} value '{text}'."

    if minimum is not None and number < minimum:
        return None, f"Row {row_number} must have a {field_name} value of at least {minimum}."

    return round(number, 2), None


def _normalize_required_text(
    value: Any,
    row_number: int,
    field_name: str,
    *,
    max_length: int,
) -> Tuple[str | None, str | None]:
    cleaned = _clean_text(value)
    if not cleaned:
        return None, f"Row {row_number} is missing '{field_name}'."

    if len(cleaned) > max_length:
        return None, f"Row {row_number} has a {field_name} value that is too long."

    return cleaned, None


def _normalize_optional_text(
    value: Any,
    row_number: int,
    field_name: str,
    *,
    max_length: int,
) -> Tuple[str | None, str | None]:
    cleaned = _clean_text(value)
    if not cleaned:
        return None, None

    if len(cleaned) > max_length:
        return None, f"Row {row_number} has a {field_name} value that is too long."

    return cleaned, None


def _normalize_grade(value: Any, row_number: int) -> Tuple[str | None, str | None]:
    cleaned = _clean_text(value)
    if not cleaned:
        return None, None

    normalized = cleaned.upper().replace(" ", "")
    if normalized == "NA":
        normalized = "N/A"

    if normalized not in VALID_GRADES:
        return None, f"Row {row_number} has an invalid grade format."

    if normalized == "P":
        normalized = "PASS"

    return normalized, None


def calculate_grade(marks: float, max_marks: float) -> str:
    """Calculate a standard letter grade from marks."""
    if max_marks <= 0:
        return "N/A"

    percentage = (marks / max_marks) * 100
    if percentage >= 90:
        return "O"
    if percentage >= 80:
        return "A+"
    if percentage >= 70:
        return "A"
    if percentage >= 60:
        return "B+"
    if percentage >= 50:
        return "B"
    if percentage >= 40:
        return "C"
    return "F"


def _validate_rows(raw_rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    cleaned_rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    seen_results: set[tuple[str, int, str]] = set()

    for row_number, raw_row in enumerate(raw_rows, start=2):
        if not any(_clean_text(value) for value in raw_row.values()):
            errors.append(f"Row {row_number} is empty.")
            continue

        register_number, register_error = _normalize_required_text(
            raw_row.get("register_number"),
            row_number,
            "register_number",
            max_length=TEXT_LIMITS["register_number"],
        )
        student_name, student_error = _normalize_required_text(
            raw_row.get("student_name"),
            row_number,
            "student_name",
            max_length=TEXT_LIMITS["student_name"],
        )
        subject_code, subject_code_error = _normalize_required_text(
            raw_row.get("subject_code"),
            row_number,
            "subject_code",
            max_length=TEXT_LIMITS["subject_code"],
        )
        semester, semester_error = _parse_semester(raw_row.get("semester"), row_number)
        marks, marks_error = _parse_number(
            raw_row.get("marks"),
            row_number,
            "marks",
            minimum=0,
        )
        max_marks, max_marks_error = _parse_number(
            raw_row.get("max_marks"),
            row_number,
            "max_marks",
            allow_blank=True,
            minimum=0.01,
        )
        subject_name, subject_name_error = _normalize_optional_text(
            raw_row.get("subject_name"),
            row_number,
            "subject_name",
            max_length=TEXT_LIMITS["subject_name"],
        )
        grade, grade_error = _normalize_grade(raw_row.get("grade"), row_number)
        campus, campus_error = _normalize_optional_text(
            raw_row.get("campus"), row_number, "campus", max_length=TEXT_LIMITS["campus"]
        )
        faculty, faculty_error = _normalize_optional_text(
            raw_row.get("faculty"), row_number, "faculty", max_length=TEXT_LIMITS["faculty"]
        )
        department, department_error = _normalize_optional_text(
            raw_row.get("department"),
            row_number,
            "department",
            max_length=TEXT_LIMITS["department"],
        )
        branch, branch_error = _normalize_optional_text(
            raw_row.get("branch"), row_number, "branch", max_length=TEXT_LIMITS["branch"]
        )
        section, section_error = _normalize_optional_text(
            raw_row.get("section"), row_number, "section", max_length=TEXT_LIMITS["section"]
        )

        row_errors = [
            error
            for error in [
                register_error,
                student_error,
                subject_code_error,
                semester_error,
                marks_error,
                max_marks_error,
                subject_name_error,
                grade_error,
                campus_error,
                faculty_error,
                department_error,
                branch_error,
                section_error,
            ]
            if error
        ]

        if row_errors:
            errors.extend(row_errors)
            continue

        register_number = register_number.upper()
        subject_code = subject_code.upper()
        subject_name = subject_name or subject_code
        max_marks = max_marks if max_marks is not None else DEFAULT_MAX_MARKS

        if marks is None or semester is None:
            errors.append(f"Row {row_number} contains incomplete marks data.")
            continue

        if marks > max_marks:
            errors.append(
                f"Row {row_number} has marks greater than max_marks ({marks} > {max_marks})."
            )
            continue

        result_key = (register_number, semester, subject_code)
        if result_key in seen_results:
            errors.append(
                f"Row {row_number} is a duplicate result entry for register number {register_number}, semester {semester}, subject {subject_code}."
            )
            continue
        seen_results.add(result_key)

        pass_status = ((marks / max_marks) * 100) >= PASS_PERCENTAGE if max_marks else False
        cleaned_rows.append(
            {
                "register_number": register_number,
                "student_name": student_name,
                "campus": campus,
                "faculty": faculty,
                "department": department,
                "branch": branch,
                "section": section,
                "semester": semester,
                "subject_code": subject_code,
                "subject_name": subject_name,
                "marks": marks,
                "max_marks": max_marks,
                "grade": grade or calculate_grade(marks, max_marks),
                "pass_status": pass_status,
            }
        )

    return cleaned_rows, errors


async def parse_upload_file(file: UploadFile) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Parse and validate an uploaded CSV or XLSX file."""
    filename = (file.filename or "upload").strip()
    lower_filename = filename.lower()

    if not lower_filename.endswith((".csv", ".xlsx")):
        return [], ["Unsupported file type. Please upload a .csv or .xlsx file."]

    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        return [], [
            f"The uploaded file is too large. The maximum allowed size is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
        ]

    if not content:
        return [], ["The uploaded file is empty."]

    if lower_filename.endswith(".csv"):
        raw_rows, errors = _read_csv_rows(content)
    else:
        raw_rows, errors = _read_xlsx_rows(content)

    if errors:
        return [], errors

    if not raw_rows:
        return [], ["The uploaded file does not contain any data rows."]

    cleaned_rows, row_errors = _validate_rows(raw_rows)
    if row_errors:
        return [], row_errors

    return cleaned_rows, []
